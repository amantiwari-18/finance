from fastapi import FastAPI, Depends, HTTPException, status, UploadFile, File
from fastapi.middleware.cors import CORSMiddleware
from fastapi.security import OAuth2PasswordRequestForm
from sqlalchemy.orm import Session
from sqlalchemy import text
from datetime import datetime, timedelta
from typing import List, Optional, Dict, Any
import pandas as pd
import io

from database import engine, get_db, Base
from models import User, TransactionType
from schemas import (
    UserCreate, UserResponse, Token, CategoryCreate, CategoryResponse, CategoryWithSubcategories,
    SubcategoryCreate, SubcategoryResponse, TransactionCreate, TransactionBulkCreate,
    TransactionResponse, TransactionWithDetails, BudgetCreate, BudgetResponse,
    RecurringTransactionCreate, RecurringTransactionResponse, AnalyticsSummary,
    CategorySummary, SubcategorySummary, LoanCreate, LoanResponse
)
from auth import get_password_hash, verify_password, create_access_token, get_current_user
from email_service import send_welcome_email, send_export_email
from scheduler import start_scheduler, stop_scheduler
from config import get_settings
import models
import openpyxl
from openpyxl.utils import get_column_letter
from fastapi.responses import StreamingResponse

# Create database tables
Base.metadata.create_all(bind=engine)

# Initialize FastAPI app
app = FastAPI(title="Expense Tracker API", version="1.0.0")

# CORS middleware
app.add_middleware(
    CORSMiddleware,
    allow_origins=["*"],
    allow_credentials=True,
    allow_methods=["*"],
    allow_headers=["*"],
)

settings = get_settings()

@app.on_event("startup")
async def startup_event():
    start_scheduler(settings)

@app.on_event("shutdown")
async def shutdown_event():
    stop_scheduler()

# ============= Authentication Endpoints =============

@app.get("/api/debug/smtp")
async def debug_smtp():
    return {
        "host": settings.SMTP_HOST,
        "port": settings.SMTP_PORT,
        "user": settings.SMTP_USER,
        "enabled": settings.SMTP_ENABLED,
        "ssl": settings.SMTP_USE_SSL
    }

@app.post("/api/auth/register", response_model=UserResponse, status_code=status.HTTP_201_CREATED)
async def register(user_data: UserCreate, db: Session = Depends(get_db)):
    """Register a new user"""
    # Check if user exists
    result = db.execute(
        text("SELECT id FROM users WHERE email = :email"),
        {"email": user_data.email}
    ).first()
    
    if result:
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail="Email already registered"
        )
    
    # Create new user
    hashed_password = get_password_hash(user_data.password)
    result = db.execute(
        text("""
            INSERT INTO users (email, hashed_password, created_at)
            VALUES (:email, :password, :created_at)
            RETURNING id, email, created_at
        """),
        {
            "email": user_data.email,
            "password": hashed_password,
            "created_at": datetime.utcnow()
        }
    )
    user_row = result.first()
    db.commit()
    
    # Refresh to get the user object if needed, but we used raw SQL so we just have a row
    user_id = user_row.id
    
    # Send welcome email
    try:
        await send_welcome_email(user_data.email)
    except Exception:
        pass # Don't fail registration if email fails
    
    # Create default categories using ORM
    from models import Category, Subcategory
    
    # Structure: (Category Name, Type, Icon, Color, [Subcategories])
    real_life_defaults = [
        ("Housing & Utilities", TransactionType.EXPENSE, "home", "#3b82f6", 
         ["Rent/Mortgage", "Electricity", "Water", "Internet", "Trash"]),
        
        ("Food & Dining", TransactionType.EXPENSE, "utensils", "#ef4444", 
         ["Groceries", "Restaurants", "Coffee Shops", "Fast Food", "Alcohol"]),
        
        ("Transportation", TransactionType.EXPENSE, "car", "#6366f1", 
         ["Fuel", "Public Transit", "Parking", "Maintenance", "Insurance"]),
        
        ("Shopping & Lifestyle", TransactionType.EXPENSE, "shopping-bag", "#ec4899", 
         ["Clothing", "Electronics", "Home Decor", "Beauty", "Pets"]),
        
        ("Health & Fitness", TransactionType.EXPENSE, "heart", "#10b981", 
         ["Gym", "Pharmacy", "Doctor", "Dental"]),
        
        ("Entertainment", TransactionType.EXPENSE, "film", "#f59e0b", 
         ["Subscriptions", "Movies", "Outings", "Hobbies"]),
        
        ("Loans & Obligations", TransactionType.EXPENSE, "landmark", "#465A65", 
         ["Student Loan", "Personal Loan", "Credit Card"]),
        
        ("Income", TransactionType.INCOME, "banknote", "#2E9CA0", 
         ["Salary", "Freelance", "Investment Returns", "Gifts/Refunds"]),
    ]
    
    try:
        for cat_name, cat_type, icon, color, subcats in real_life_defaults:
            # Create Category
            new_cat = Category(
                name=cat_name,
                type=cat_type,
                icon=icon,
                color=color,
                user_id=user_id,
                created_at=datetime.utcnow()
            )
            db.add(new_cat)
            db.flush() # Get the category ID
            
            # Create Subcategories for this category
            for sub_name in subcats:
                new_sub = Subcategory(
                    name=sub_name,
                    category_id=new_cat.id,
                    user_id=user_id,
                    created_at=datetime.utcnow()
                )
                db.add(new_sub)
        
        db.commit()
    except Exception as e:
        db.rollback()
        # Log error but don't fail registration as the user is already created
        print(f"Error creating default categories for {user_data.email}: {str(e)}")
    
    return {"id": user_row.id, "email": user_row.email, "created_at": user_row.created_at}

@app.post("/api/auth/login", response_model=Token)
async def login(form_data: OAuth2PasswordRequestForm = Depends(), db: Session = Depends(get_db)):
    """Login and get access token"""
    result = db.execute(
        text("SELECT id, email, hashed_password FROM users WHERE email = :email"),
        {"email": form_data.username}
    ).first()
    
    if not result or not verify_password(form_data.password, result.hashed_password):
        raise HTTPException(
            status_code=status.HTTP_401_UNAUTHORIZED,
            detail="Incorrect email or password",
            headers={"WWW-Authenticate": "Bearer"},
        )
    
    access_token = create_access_token(data={"sub": result.email})
    return {"access_token": access_token, "token_type": "bearer"}

# ============= Category Endpoints =============

@app.get("/api/categories", response_model=List[CategoryWithSubcategories])
async def get_categories(
    current_user: User = Depends(get_current_user),
    db: Session = Depends(get_db)
):
    """Get all categories with subcategories"""
    from models import Category, Subcategory
    
    categories = db.query(Category).filter(Category.user_id == current_user.id).all()
    subcategories = db.query(Subcategory).filter(Subcategory.user_id == current_user.id).all()
    
    # Group subcategories by category_id
    subcats_by_category = {}
    for sub in subcategories:
        if sub.category_id not in subcats_by_category:
            subcats_by_category[sub.category_id] = []
        subcats_by_category[sub.category_id].append(sub)
    
    result = []
    for cat in categories:
        result.append({
            "id": cat.id,
            "name": cat.name,
            "type": cat.type,
            "color": cat.color,
            "icon": cat.icon,
            "created_at": cat.created_at,
            "subcategories": subcats_by_category.get(cat.id, [])
        })
    
    return result

@app.post("/api/categories", response_model=CategoryResponse, status_code=status.HTTP_201_CREATED)
async def create_category(
    category_data: CategoryCreate,
    current_user: User = Depends(get_current_user),
    db: Session = Depends(get_db)
):
    """Create a new category"""
    from models import Category
    db_category = Category(
        name=category_data.name,
        type=category_data.type,
        icon=category_data.icon or "shopping-bag",
        color=category_data.color or "#6366f1",
        user_id=current_user.id,
        created_at=datetime.utcnow()
    )
    db.add(db_category)
    try:
        db.commit()
        db.refresh(db_category)
    except Exception as e:
        db.rollback()
        raise HTTPException(status_code=500, detail=str(e))
    
    return db_category

@app.delete("/api/categories/{category_id}", status_code=status.HTTP_204_NO_CONTENT)
async def delete_category(
    category_id: int,
    current_user: User = Depends(get_current_user),
    db: Session = Depends(get_db)
):
    """Delete a category and cascade-delete its subcategories and transactions"""
    try:
        # 1. Delete transactions in this category (FK: transactions.category_id)
        db.execute(
            text("DELETE FROM transactions WHERE category_id = :cat_id AND user_id = :user_id"),
            {"cat_id": category_id, "user_id": current_user.id}
        )
        # 2. Delete subcategories in this category (FK: subcategories.category_id)
        db.execute(
            text("DELETE FROM subcategories WHERE category_id = :cat_id AND user_id = :user_id"),
            {"cat_id": category_id, "user_id": current_user.id}
        )
        # 3. Delete the category itself
        result = db.execute(
            text("DELETE FROM categories WHERE id = :id AND user_id = :user_id"),
            {"id": category_id, "user_id": current_user.id}
        )
        db.commit()
        if result.rowcount == 0:
            raise HTTPException(status_code=404, detail="Category not found")
    except HTTPException:
        raise
    except Exception as e:
        db.rollback()
        raise HTTPException(status_code=500, detail=str(e))
    return None

# ============= Subcategory Endpoints =============

@app.get("/api/subcategories", response_model=List[SubcategoryResponse])
async def get_subcategories(
    category_id: Optional[int] = None,
    current_user: User = Depends(get_current_user),
    db: Session = Depends(get_db)
):
    """Get subcategories, optionally filtered by category"""
    if category_id:
        result = db.execute(
            text("""
                SELECT id, name, category_id, created_at
                FROM subcategories
                WHERE user_id = :user_id AND category_id = :category_id
                ORDER BY created_at
            """),
            {"user_id": current_user.id, "category_id": category_id}
        )
    else:
        result = db.execute(
            text("""
                SELECT id, name, category_id, created_at
                FROM subcategories
                WHERE user_id = :user_id
                ORDER BY created_at
            """),
            {"user_id": current_user.id}
        )
    
    subcategories = result.fetchall()
    return [
        {
            "id": sub.id,
            "name": sub.name,
            "category_id": sub.category_id,
            "created_at": sub.created_at
        }
        for sub in subcategories
    ]

@app.post("/api/subcategories", response_model=SubcategoryResponse, status_code=status.HTTP_201_CREATED)
async def create_subcategory(
    subcategory_data: SubcategoryCreate,
    current_user: User = Depends(get_current_user),
    db: Session = Depends(get_db)
):
    """Create a new subcategory"""
    # Verify category belongs to user
    from models import Category, Subcategory
    category_check = db.query(Category).filter(
        Category.id == subcategory_data.category_id,
        Category.user_id == current_user.id
    ).first()
    
    if not category_check:
        raise HTTPException(status_code=404, detail="Category not found")
    
    db_subcategory = Subcategory(
        name=subcategory_data.name,
        category_id=subcategory_data.category_id,
        user_id=current_user.id,
        created_at=datetime.utcnow()
    )
    db.add(db_subcategory)
    try:
        db.commit()
        db.refresh(db_subcategory)
    except Exception as e:
        db.rollback()
        raise HTTPException(status_code=500, detail=str(e))
    
    return db_subcategory

@app.delete("/api/subcategories/{subcategory_id}", status_code=status.HTTP_204_NO_CONTENT)
async def delete_subcategory(
    subcategory_id: int,
    current_user: User = Depends(get_current_user),
    db: Session = Depends(get_db)
):
    """Delete a subcategory"""
    result = db.execute(
        text("DELETE FROM subcategories WHERE id = :id AND user_id = :user_id"),
        {"id": subcategory_id, "user_id": current_user.id}
    )
    db.commit()
    if result.rowcount == 0:
        raise HTTPException(status_code=404, detail="Subcategory not found")
    return None

# ============= Transaction Endpoints =============

@app.get("/api/transactions", response_model=List[TransactionWithDetails])
async def get_transactions(
    category_id: Optional[int] = None,
    subcategory_id: Optional[int] = None,
    start_date: Optional[datetime] = None,
    end_date: Optional[datetime] = None,
    transaction_type: Optional[TransactionType] = None,
    skip: int = 0,
    limit: int = 100,
    current_user: User = Depends(get_current_user),
    db: Session = Depends(get_db)
):
    """Get transactions with filters"""
    from models import Transaction, Category, Subcategory
    
    query = db.query(Transaction).join(Category).join(Subcategory).filter(Transaction.user_id == current_user.id)
    
    if category_id:
        query = query.filter(Transaction.category_id == category_id)
    if subcategory_id:
        query = query.filter(Transaction.subcategory_id == subcategory_id)
    if start_date:
        query = query.filter(Transaction.date >= start_date)
    if end_date:
        query = query.filter(Transaction.date <= end_date)
    if transaction_type:
        query = query.filter(Category.type == transaction_type)
    
    transactions = query.order_by(Transaction.date.desc()).offset(skip).limit(limit).all()
    
    return transactions

@app.post("/api/transactions", response_model=TransactionResponse, status_code=status.HTTP_201_CREATED)
async def create_transaction(
    transaction_data: TransactionCreate,
    current_user: User = Depends(get_current_user),
    db: Session = Depends(get_db)
):
    """Create a single transaction"""
    # Verify category and subcategory
    from models import Category, Subcategory, Transaction
    category_check = db.query(Category).filter(
        Category.id == transaction_data.category_id,
        Category.user_id == current_user.id
    ).first()
    
    if not category_check:
        raise HTTPException(status_code=404, detail="Category not found")
    
    subcategory_check = db.query(Subcategory).filter(
        Subcategory.id == transaction_data.subcategory_id,
        Subcategory.user_id == current_user.id
    ).first()
    
    if not subcategory_check:
        raise HTTPException(status_code=404, detail="Subcategory not found")
    
    db_transaction = Transaction(
        user_id=current_user.id,
        category_id=transaction_data.category_id,
        subcategory_id=transaction_data.subcategory_id,
        amount=transaction_data.amount,
        date=transaction_data.date,
        description=transaction_data.description,
        receipt_url=transaction_data.receipt_url,
        created_at=datetime.utcnow()
    )
    db.add(db_transaction)
    
    # Update Loan balance if subcategory is linked to a loan
    from models import Loan
    loan = db.query(Loan).filter(Loan.subcategory_id == transaction_data.subcategory_id).first()
    if loan:
        loan.remaining_amount -= transaction_data.amount

    try:
        db.commit()
        db.refresh(db_transaction)
    except Exception as e:
        db.rollback()
        raise HTTPException(status_code=500, detail=str(e))
    
    return db_transaction

@app.post("/api/transactions/bulk", status_code=status.HTTP_201_CREATED)
async def create_bulk_transactions(
    bulk_data: TransactionBulkCreate,
    current_user: User = Depends(get_current_user),
    db: Session = Depends(get_db)
):
    """Create multiple transactions at once"""
    created_count = 0
    
    for transaction_data in bulk_data.transactions:
        # Verify category and subcategory
        category_check = db.execute(
            text("SELECT id FROM categories WHERE id = :id AND user_id = :user_id"),
            {"id": transaction_data.category_id, "user_id": current_user.id}
        ).first()
        
        subcategory_check = db.execute(
            text("SELECT id FROM subcategories WHERE id = :id AND user_id = :user_id"),
            {"id": transaction_data.subcategory_id, "user_id": current_user.id}
        ).first()
        
        if category_check and subcategory_check:
            db.execute(
                text("""
                    INSERT INTO transactions (user_id, category_id, subcategory_id, amount, date, description, receipt_url, created_at)
                    VALUES (:user_id, :category_id, :subcategory_id, :amount, :date, :description, :receipt_url, :created_at)
                """),
                {
                    "user_id": current_user.id,
                    "category_id": transaction_data.category_id,
                    "subcategory_id": transaction_data.subcategory_id,
                    "amount": transaction_data.amount,
                    "date": transaction_data.date,
                    "description": transaction_data.description,
                    "receipt_url": transaction_data.receipt_url,
                    "created_at": datetime.utcnow()
                }
            )
            created_count += 1
    
    db.commit()
    return {"created": created_count, "message": f"Successfully created {created_count} transactions"}

@app.put("/api/transactions/{transaction_id}", response_model=TransactionResponse)
async def update_transaction(
    transaction_id: int,
    transaction_data: TransactionCreate,
    current_user: User = Depends(get_current_user),
    db: Session = Depends(get_db)
):
    """Update a transaction"""
    from models import Transaction
    db_transaction = db.query(Transaction).filter(
        Transaction.id == transaction_id,
        Transaction.user_id == current_user.id
    ).first()
    
    if not db_transaction:
        raise HTTPException(status_code=404, detail="Transaction not found")
    
    db_transaction.category_id = transaction_data.category_id
    db_transaction.subcategory_id = transaction_data.subcategory_id
    db_transaction.amount = transaction_data.amount
    db_transaction.date = transaction_data.date
    db_transaction.description = transaction_data.description
    db_transaction.receipt_url = transaction_data.receipt_url
    
    try:
        db.commit()
        db.refresh(db_transaction)
    except Exception as e:
        db.rollback()
        raise HTTPException(status_code=500, detail=str(e))
    
    return db_transaction

@app.delete("/api/transactions/{transaction_id}", status_code=status.HTTP_204_NO_CONTENT)
async def delete_transaction(
    transaction_id: int,
    current_user: User = Depends(get_current_user),
    db: Session = Depends(get_db)
):
    """Delete a transaction"""
    result = db.execute(
        text("DELETE FROM transactions WHERE id = :id AND user_id = :user_id"),
        {"id": transaction_id, "user_id": current_user.id}
    )
    db.commit()
    
    if result.rowcount == 0:
        raise HTTPException(status_code=404, detail="Transaction not found")
    
    return None

@app.get("/api/export/transactions")
async def export_transactions(
    range_type: str = "all", # today, week, month, year, all
    method: str = "download", # download, email
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user)
):
    """Export all transactions to a professional Excel report"""
    import pandas as pd
    import io
    from fastapi.responses import StreamingResponse
    from datetime import datetime, timedelta
    from openpyxl.utils import get_column_letter
    import models

    # Calculate date filter
    now = datetime.now()
    filters = [models.Transaction.user_id == current_user.id]
    
    if range_type == "today":
        filters.append(models.Transaction.date >= now.replace(hour=0, minute=0, second=0, microsecond=0))
    elif range_type == "week":
        filters.append(models.Transaction.date >= now - timedelta(days=7))
    elif range_type == "month":
        filters.append(models.Transaction.date >= now - timedelta(days=30))
    elif range_type == "year":
        filters.append(models.Transaction.date >= now - timedelta(days=365))

    # Fetch data joined with categories for type info
    txs = db.query(models.Transaction).filter(*filters).order_by(models.Transaction.date.desc()).all()
    
    data = []
    for t in txs:
        data.append({
            "Date": t.date.strftime("%Y-%m-%d %H:%M"),
            "Flow Type": t.category.type.value.capitalize(),
            "Category": t.category.name,
            "Subcategory": t.subcategory.name if t.subcategory else "-",
            "Amount": float(t.amount),
            "Currency": "INR",
            "Description": t.description or "-"
        })

    if not data:
        if method == "email":
             return {"status": "info", "message": "No records found for this period to email."}
        df = pd.DataFrame(columns=["Date", "Flow Type", "Category", "Subcategory", "Amount", "Currency", "Description"])
    else:
        df = pd.DataFrame(data)
    
    # Create colorful Excel using openpyxl
    output = io.BytesIO()
    with pd.ExcelWriter(output, engine='openpyxl') as writer:
        df.to_excel(writer, index=False, sheet_name='Vault_Registry')
        workbook = writer.book
        worksheet = writer.sheets['Vault_Registry']
        
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        
        header_fill = PatternFill(start_color='6366F1', end_color='6366F1', fill_type='solid')
        header_font = Font(color='FFFFFF', bold=True, size=12)
        center_align = Alignment(horizontal='center', vertical='center')
        border = Border(left=Side(style='thin', color='CBD5E1'), 
                        right=Side(style='thin', color='CBD5E1'), 
                        top=Side(style='thin', color='CBD5E1'), 
                        bottom=Side(style='thin', color='CBD5E1'))

        # Header styling and Auto-align
        for col_num, value in enumerate(df.columns.values):
            cell = worksheet.cell(row=1, column=col_num + 1)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = center_align
            
            # Auto-align width
            max_length = len(str(value)) + 2
            col_letter = get_column_letter(col_num + 1)
            for cell_in_col in worksheet[col_letter]:
                try:
                    if len(str(cell_in_col.value)) > max_length:
                        max_length = len(str(cell_in_col.value))
                except: pass
            worksheet.column_dimensions[col_letter].width = min(max_length + 2, 50)

        # Data styling (alternating shading)
        even_fill = PatternFill(start_color='F8FAFC', end_color='F8FAFC', fill_type='solid')
        for row_num in range(2, len(df) + 2):
            fill = even_fill if row_num % 2 == 0 else None
            for col_num in range(1, len(df.columns) + 1):
                cell = worksheet.cell(row=row_num, column=col_num)
                if fill: cell.fill = fill
                cell.border = border
                if df.columns[col_num-1] == 'Amount':
                    cell.number_format = '₹#,##0'

    output.seek(0)
    excel_data = output.read()

    if method == "email":
        attach_excel = range_type not in ["today", "week"]
        results_for_email = []
        for d in data:
            results_for_email.append({
                "category": d["Category"],
                "description": d["Description"],
                "amount": d["Amount"],
                "flow_type": d["Flow Type"].lower()
            })
            
        import email_service
        try:
            await email_service.send_export_email(
                email=current_user.email,
                range_name=range_type,
                results=results_for_email,
                attachment=excel_data if (attach_excel or not results_for_email) else None,
                filename=f"Vault_Registry_{range_type}.xlsx"
            )
            return {"status": "success", "message": f"Export email sent for range: {range_type}"}
        except Exception as e:
            raise HTTPException(
                status_code=status.HTTP_500_INTERNAL_SERVER_ERROR,
                detail=f"Vault Transmission Error: {str(e)}"
            )
    
    else: # method == "download"
        return StreamingResponse(
            io.BytesIO(excel_data),
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={"Content-Disposition": f"attachment; filename=Vault_Registry_{range_type}.xlsx"}
        )


# ============= Budget Endpoints =============

@app.get("/api/budgets", response_model=List[BudgetResponse])
async def get_budgets(
    current_user: User = Depends(get_current_user),
    db: Session = Depends(get_db)
):
    """Get all budgets"""
    from models import Budget
    budgets = db.query(Budget).filter(Budget.user_id == current_user.id).order_by(Budget.start_date.desc()).all()
    return budgets

@app.post("/api/budgets", response_model=BudgetResponse, status_code=status.HTTP_201_CREATED)
async def create_budget(
    budget_data: BudgetCreate,
    current_user: User = Depends(get_current_user),
    db: Session = Depends(get_db)
):
    """Create a new budget"""
    from models import Budget
    db_budget = Budget(
        user_id=current_user.id,
        category_id=budget_data.category_id,
        subcategory_id=budget_data.subcategory_id,
        amount=budget_data.amount,
        period=budget_data.period,
        start_date=datetime.utcnow()
    )
    db.add(db_budget)
    try:
        db.commit()
        db.refresh(db_budget)
    except Exception as e:
        db.rollback()
        raise HTTPException(status_code=500, detail=str(e))
    
    return db_budget

# ============= Loan Endpoints =============

@app.get("/api/loans", response_model=List[LoanResponse])
async def get_loans(
    current_user: User = Depends(get_current_user),
    db: Session = Depends(get_db)
):
    """Get all loans"""
    from models import Loan
    return db.query(Loan).filter(Loan.user_id == current_user.id).all()

@app.post("/api/loans", response_model=LoanResponse, status_code=status.HTTP_201_CREATED)
async def create_loan(
    loan_data: LoanCreate,
    current_user: User = Depends(get_current_user),
    db: Session = Depends(get_db)
):
    """Create a new loan and its tracking subcategory"""
    from models import Category, Subcategory, Loan
    
    # 1. Find or create "Loans & EMI" category
    loan_cat = db.query(Category).filter(
        Category.name == "Loans & EMI", 
        Category.user_id == current_user.id
    ).first()
    
    if not loan_cat:
        loan_cat = Category(
            name="Loans & EMI",
            user_id=current_user.id,
            type=TransactionType.EXPENSE,
            icon="landmark",
            color="#ef4444"
        )
        db.add(loan_cat)
        db.commit()
        db.refresh(loan_cat)
    
    # 2. Create subcategory for this specific loan
    sub = Subcategory(
        name=loan_data.name,
        category_id=loan_cat.id,
        user_id=current_user.id
    )
    db.add(sub)
    db.commit()
    db.refresh(sub)
    
    # 3. Create the Loan
    db_loan = Loan(
        user_id=current_user.id,
        name=loan_data.name,
        total_amount=loan_data.total_amount,
        remaining_amount=loan_data.total_amount,
        interest_rate=loan_data.interest_rate,
        subcategory_id=sub.id
    )
    db.add(db_loan)
    db.commit()
    db.refresh(db_loan)
    
    return db_loan

@app.delete("/api/loans/{loan_id}", status_code=status.HTTP_204_NO_CONTENT)
async def delete_loan(
    loan_id: int,
    current_user: User = Depends(get_current_user),
    db: Session = Depends(get_db)
):
    """Delete a loan"""
    from models import Loan
    db.query(Loan).filter(Loan.id == loan_id, Loan.user_id == current_user.id).delete()
    db.commit()
    return None

# ============= Analytics Endpoints =============

@app.get("/api/analytics/summary", response_model=AnalyticsSummary)
async def get_analytics_summary(
    start_date: Optional[datetime] = None,
    end_date: Optional[datetime] = None,
    current_user: User = Depends(get_current_user),
    db: Session = Depends(get_db)
):
    """Get spending summary with category and subcategory breakdown"""
    from models import Transaction, Category, Subcategory
    from sqlalchemy import func
    
    try:
        # 1. Total Income and Expense
        type_query = db.query(
            Category.type.label('type'), 
            func.sum(Transaction.amount).label('total')
        ).join(Transaction).filter(Transaction.user_id == current_user.id)
        
        if start_date:
            type_query = type_query.filter(Transaction.date >= start_date)
        if end_date:
            type_query = type_query.filter(Transaction.date <= end_date)
        
        type_results = type_query.group_by(Category.type).all()
        
        total_income = 0.0
        total_expense = 0.0
        for row in type_results:
            if row.type == "income":
                total_income = float(row.total or 0)
            elif row.type == "expense":
                total_expense = float(row.total or 0)
        
        # 2. Category breakdown
        cat_query = db.query(
            Category.id.label('category_id'), 
            Category.name.label('category_name'),
            func.sum(Transaction.amount).label('total'),
            func.count(Transaction.id).label('count')
        ).join(Transaction).filter(Transaction.user_id == current_user.id)
        
        if start_date:
            cat_query = cat_query.filter(Transaction.date >= start_date)
        if end_date:
            cat_query = cat_query.filter(Transaction.date <= end_date)
            
        cat_results = cat_query.group_by(Category.id, Category.name).all()
        
        # Fetch budgets to map them
        from models import Budget
        budgets = db.query(Budget).filter(Budget.user_id == current_user.id).all()
        budget_map = {b.category_id: b.amount for b in budgets if b.category_id}

        category_breakdown = []
        for row in cat_results:
            category_breakdown.append({
                "category_id": row.category_id,
                "category_name": row.category_name,
                "total_amount": float(row.total or 0),
                "transaction_count": row.count,
                "budget_amount": budget_map.get(row.category_id, 0.0)
            })
        
        # 3. Subcategory breakdown
        sub_query = db.query(
            Subcategory.id.label('subcategory_id'),
            Subcategory.name.label('subcategory_name'),
            Category.name.label('category_name'),
            func.sum(Transaction.amount).label('total'),
            func.count(Transaction.id).label('count')
        ).select_from(Subcategory)\
         .join(Transaction, Subcategory.id == Transaction.subcategory_id)\
         .join(Category, Transaction.category_id == Category.id)\
         .filter(Transaction.user_id == current_user.id)
        
        if start_date:
            sub_query = sub_query.filter(Transaction.date >= start_date)
        if end_date:
            sub_query = sub_query.filter(Transaction.date <= end_date)
            
        sub_results = sub_query.group_by(Subcategory.id, Subcategory.name, Category.name).all()
        subcategory_breakdown = []
        for row in sub_results:
            subcategory_breakdown.append({
                "subcategory_id": row.subcategory_id,
                "subcategory_name": row.subcategory_name,
                "category_name": row.category_name,
                "total_amount": float(row.total or 0),
                "transaction_count": row.count
            })
        
        return {
            "total_income": total_income,
            "total_expense": total_expense,
            "balance": total_income - total_expense,
            "category_breakdown": category_breakdown,
            "subcategory_breakdown": subcategory_breakdown
        }
    except Exception as e:
        import traceback
        traceback.print_exc()
        raise HTTPException(status_code=500, detail=str(e))

# ============= Import Endpoint =============

@app.post("/api/import/excel")
async def import_excel(
    file: UploadFile = File(...),
    current_user: User = Depends(get_current_user),
    db: Session = Depends(get_db)
):
    """Import transactions from Excel file"""
    try:
        contents = await file.read()
        df = pd.read_excel(io.BytesIO(contents))
        
        required_columns = ['Date', 'Category', 'Subcategory', 'Amount']
        if not all(col in df.columns for col in required_columns):
            raise HTTPException(
                status_code=400,
                detail=f"Excel file must contain columns: {', '.join(required_columns)}"
            )
        
        created_count = 0
        
        for _, row in df.iterrows():
            # Find or create category
            category = db.execute(
                text("SELECT id FROM categories WHERE name = :name AND user_id = :user_id"),
                {"name": row['Category'], "user_id": current_user.id}
            ).first()
            
            if not category:
                category = db.execute(
                    text("""
                        INSERT INTO categories (name, type, user_id, created_at)
                        VALUES (:name, :type, :user_id, :created_at)
                        RETURNING id
                    """),
                    {
                        "name": row['Category'],
                        "type": "expense",
                        "user_id": current_user.id,
                        "created_at": datetime.utcnow()
                    }
                ).first()
            
            # Find or create subcategory
            subcategory = db.execute(
                text("SELECT id FROM subcategories WHERE name = :name AND category_id = :category_id AND user_id = :user_id"),
                {"name": row['Subcategory'], "category_id": category.id, "user_id": current_user.id}
            ).first()
            
            if not subcategory:
                subcategory = db.execute(
                    text("""
                        INSERT INTO subcategories (name, category_id, user_id, created_at)
                        VALUES (:name, :category_id, :user_id, :created_at)
                        RETURNING id
                    """),
                    {
                        "name": row['Subcategory'],
                        "category_id": category.id,
                        "user_id": current_user.id,
                        "created_at": datetime.utcnow()
                    }
                ).first()
            
            # Create transaction
            db.execute(
                text("""
                    INSERT INTO transactions (user_id, category_id, subcategory_id, amount, date, description, created_at)
                    VALUES (:user_id, :category_id, :subcategory_id, :amount, :date, :description, :created_at)
                """),
                {
                    "user_id": current_user.id,
                    "category_id": category.id,
                    "subcategory_id": subcategory.id,
                    "amount": float(row['Amount']),
                    "date": pd.to_datetime(row['Date']),
                    "description": row.get('Description', None),
                    "created_at": datetime.utcnow()
                }
            )
            created_count += 1
        
        db.commit()
        return {"message": f"Successfully imported {created_count} transactions"}
    
    except Exception as e:
        raise HTTPException(status_code=400, detail=f"Error importing file: {str(e)}")

@app.get("/")
async def root():
    """Root endpoint"""
    return {"message": "Expense Tracker API", "docs": "/docs"}
